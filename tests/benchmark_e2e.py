"""
benchmark_e2e.py
================
End-to-End (E2E) LLM Benchmark Runner — 3 Asama, 3 Sheet Excel.

Calistirma sirasi her config ID icin:
  1. OT-128  — Kripto OT 128-bit guvenlikle calistirilir
  2. OT-256  — Kripto OT 256-bit guvenlikle calistirilir
  3. Plain   — Kriptosuz plain bitwise-AND intersection

Sonuclar tek bir Excel dosyasina yazilir (tests/results/sample_output.xlsx):
  Sheet 1: "OT-128 Benchmark"   — config + agent_a/b array + sure (saniye) + intersection
  Sheet 2: "OT-256 Benchmark"   — ayni yapi
  Sheet 3: "Plain Intersection" — config + agent_a/b array + sure (saniye) + intersection
                                  (Guvenlik Bit sutunu YOK — plain bit kullanmaz)

Kullanim
--------
    cd c:/Users/lenovo/Terrarium
    python tests/benchmark_e2e.py --max-rows 5
    python tests/benchmark_e2e.py            # tum 143 config
    python tests/benchmark_e2e.py --no-ot    # sadece plain
"""

from __future__ import annotations

import array
import ast
import asyncio
import csv
import logging
import re
import sys
import time
import argparse
from dataclasses import dataclass, field
from pathlib import Path
from statistics import mean, stdev
from typing import List, Optional

# ---------------------------------------------------------------------------
# PATH SETUP
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent    # …/Terrarium
TESTS_DIR = Path(__file__).resolve().parent           # …/Terrarium/tests
CRYPTO_DIR = REPO_ROOT / "crypto"

for p in (str(REPO_ROOT), str(TESTS_DIR), str(CRYPTO_DIR)):
    if p not in sys.path:
        sys.path.insert(0, p)

# ---------------------------------------------------------------------------
# OPTIONAL IMPORTS
# ---------------------------------------------------------------------------

# OT module
OT_AVAILABLE = False
OTManager = None
try:
    from crypto.ot_manager import OTManager as _OTM   # type: ignore
    OTManager = _OTM
    OT_AVAILABLE = True
except Exception as e:
    logging.warning("OT modulu yuklenemedi: %s", e)

# openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Terrarium codebase (E2E icin)
E2E_AVAILABLE = False
try:
    from examples.base_main import run_simulation          # type: ignore
    from src.utils import load_config                      # type: ignore
    from envs.dcops.meeting_scheduling.meeting_scheduling_env import (  # type: ignore
        MeetingSchedulingEnvironment,
    )
    import crypto as _crypto                               # type: ignore
    E2E_AVAILABLE = True
except Exception as e:
    logging.warning("Terrarium E2E modulleri yuklenemedi: %s — E2E modu devre disi.", e)

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------
logging.basicConfig(
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PATHS
# ---------------------------------------------------------------------------
DEFAULT_CSV = (
    REPO_ROOT
    / "examples" / "outputs" / "availability_configurations"
    / "availability_configurations.csv"
)
DEFAULT_OUT = TESTS_DIR / "results" / "sample_output.xlsx"

# ---------------------------------------------------------------------------
# PLAIN INTERSECTION  (array modulu ile)
# ---------------------------------------------------------------------------

def plain_intersection(agent_a: List[int], agent_b: List[int]) -> array.array:
    """Bitwise AND: her iki agentta da 1 olan slot indekslerini doner."""
    if len(agent_a) != len(agent_b):
        raise ValueError(f"Uzunluklar esit degil: {len(agent_a)} != {len(agent_b)}")
    result = array.array('i')
    for idx, (a, b) in enumerate(zip(agent_a, agent_b)):
        if a == 1 and b == 1:
            result.append(idx)
    return result

# ---------------------------------------------------------------------------
# DATA STRUCTURES
# ---------------------------------------------------------------------------

@dataclass
class ConfigRow:
    config_id: int
    negotiation_window_days: int
    time_slot_minutes: int
    security_bits: int
    availability_density: float
    intersection_density: float
    total_slots: int
    actual_intersections: int
    agent_a: List[int]
    agent_b: List[int]


@dataclass
class RunResult:
    """Bir config x mod icin tek calisma sonucu."""
    config_id: int
    total_slots: int
    security_bits: int          # OT icin 128/256; plain icin 0
    availability_density: float
    intersection_density: float
    actual_intersections: int
    mode: str                   # "OT-128" | "OT-256" | "PLAIN"
    agent_a: List[int] = field(default_factory=list)
    agent_b: List[int] = field(default_factory=list)
    # Sureleri saniye cinsinden
    duration_s: float = 0.0
    intersection_array: str = ""
    intersection_size: int = 0
    success: bool = False
    error: str = ""

# ---------------------------------------------------------------------------
# CSV PARSER
# ---------------------------------------------------------------------------
_ARRAY_RE = re.compile(r'"?\[([^\]]+)\]"?')


def load_configurations(csv_path: Path) -> List[ConfigRow]:
    configs: List[ConfigRow] = []
    with open(csv_path, newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        next(reader, None)   # header
        for raw_row in reader:
            full = ",".join(raw_row).strip()
            arrays = _ARRAY_RE.findall(full)
            if len(arrays) < 2:
                continue
            try:
                agent_a = [int(x.strip()) for x in arrays[-2].split(",")]
                agent_b = [int(x.strip()) for x in arrays[-1].split(",")]
            except ValueError:
                continue
            scalar_part = _ARRAY_RE.sub("", full).replace('"', "")
            scalar_part = re.sub(r",+$", "", scalar_part.strip())
            meta = scalar_part.split(",")
            try:
                configs.append(ConfigRow(
                    config_id=int(meta[0]),
                    negotiation_window_days=int(meta[1]),
                    time_slot_minutes=int(meta[2]),
                    security_bits=int(meta[3]),
                    availability_density=float(meta[4]),
                    intersection_density=float(meta[5]),
                    total_slots=int(meta[7]),
                    actual_intersections=int(meta[11]),
                    agent_a=agent_a,
                    agent_b=agent_b,
                ))
            except (IndexError, ValueError):
                continue
    logger.info("%d konfigürasyon yüklendi: %s", len(configs), csv_path)
    return configs

# ---------------------------------------------------------------------------
# MONKEY PATCHING  (ana kodu degistirmeden runtime override)
# ---------------------------------------------------------------------------

def _patch_environment(agent_a: List[int], agent_b: List[int]) -> None:
    """LLM ortamini CSV array'leriyle besler."""
    if not E2E_AVAILABLE:
        return

    def _mocked_availability(self, participants: list):
        av = {}
        for p in participants:
            av[p] = agent_a if "A" in p else agent_b
        return av

    MeetingSchedulingEnvironment._generate_simulated_availability = _mocked_availability


def _patch_crypto(mode: str, bit_size: int) -> None:
    """
    mode='OT'   → OT manager ile (bit_size=128 veya 256)
    mode='PLAIN' → plain bitwise intersection ile
    """
    if not E2E_AVAILABLE:
        return

    if mode == "OT":
        def _ot_fn(sender, receiver, total_slots=12):
            mgr = OTManager(bit_size=bit_size)
            return mgr.compute_intersection(sender, receiver, total_slots)
        _crypto.compute_private_intersection = _ot_fn
    else:
        def _plain_fn(sender, receiver, total_slots=12):
            return plain_intersection(sender, receiver)
        _crypto.compute_private_intersection = _plain_fn

# ---------------------------------------------------------------------------
# EXCEL WRITER  (3 sheet)
# ---------------------------------------------------------------------------

def _hdr(cell, bg: str = "2E4057") -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _sub(cell, bg: str = "4A6FA5") -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, max_w: int = 80) -> None:
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[ltr].width = min(mx + 2, max_w)


def _write_ot_sheet(ws, results: List[RunResult], bit: int) -> None:
    """OT-128 veya OT-256 sheet yazar."""
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    # Grup basliklar (satir 1)
    groups = [
        ("Config Bilgisi",       6),
        ("Girdi Arrayleri",      2),
        (f"OT-{bit} Sonuclari",  4),
        ("Durum",                1),
    ]
    ot_color = "0D3B66" if bit == 128 else "023E8A"
    colors = ["2E4057", "1B4332", ot_color, "6B2737"]
    col = 1
    for (label, span), color in zip(groups, colors):
        c = ws.cell(row=1, column=col, value=label)
        _hdr(c, bg=color)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + span - 1)
        col += span

    # Sutun basliklar (satir 2)
    cols = [
        "Config ID", "Toplam Slot", f"OT Bit ({bit})", "Musaitlik Yogunlugu",
        "Kesisim Yogunlugu", "Gercek Kesisim",
        "Agent A Array", "Agent B Array",
        f"OT-{bit} Sure (saniye)", f"OT-{bit} Kesisim Arrayi",
        f"OT-{bit} Boyut", "E2E Basarili?",
        "Hata",
    ]
    for c, name in enumerate(cols, 1):
        _sub(ws.cell(row=2, column=c, value=name), bg="4A6FA5")

    # Veri satirlari
    ot_results = [r for r in results if r.mode == f"OT-{bit}"]
    for ri, r in enumerate(ot_results, 3):
        row_data = [
            r.config_id, r.total_slots, bit,
            r.availability_density, r.intersection_density, r.actual_intersections,
            str(r.agent_a), str(r.agent_b),
            r.duration_s, r.intersection_array, r.intersection_size,
            "EVET" if r.success else "HAYIR",
            r.error,
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top")
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF4FA")

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _write_plain_sheet(ws, results: List[RunResult]) -> None:
    """Plain Intersection sheet yazar — Guvenlik Bit sutunu YOK."""
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    groups = [
        ("Config Bilgisi - Plain", 5),   # Bit yok → 5 sutun
        ("Girdi Arrayleri",        2),
        ("Plain Sonuclari",        3),
        ("Durum",                  1),
    ]
    colors = ["1B4332", "1A6B40", "40916C", "145A32"]
    col = 1
    for (label, span), color in zip(groups, colors):
        c = ws.cell(row=1, column=col, value=label)
        _hdr(c, bg=color)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + span - 1)
        col += span

    # Sutun basliklar — Guvenlik Bit YOK
    cols = [
        "Config ID", "Toplam Slot", "Musaitlik Yogunlugu",
        "Kesisim Yogunlugu", "Gercek Kesisim",
        "Agent A Array", "Agent B Array",
        "Plain Sure (saniye)", "Plain Kesisim Arrayi", "Plain Boyut",
        "Hata",
    ]
    for c, name in enumerate(cols, 1):
        _sub(ws.cell(row=2, column=c, value=name), bg="40916C")

    plain_results = [r for r in results if r.mode == "PLAIN"]
    for ri, r in enumerate(plain_results, 3):
        row_data = [
            r.config_id, r.total_slots,          # security_bits YOK
            r.availability_density, r.intersection_density, r.actual_intersections,
            str(r.agent_a), str(r.agent_b),
            r.duration_s, r.intersection_array, r.intersection_size,
            r.error,
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top")
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="E8F8F0")

    ws.freeze_panes = "A3"
    _auto_width(ws)


def write_excel(results: List[RunResult], out_path: Path) -> None:
    """3 sheet'li Excel yazar: OT-128 | OT-256 | Plain Intersection."""
    if not EXCEL_AVAILABLE:
        print("UYARI: openpyxl yuklu degil. 'pip install openpyxl' calistirin.")
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws128 = wb.active
    ws128.title = "OT-128 Benchmark"
    _write_ot_sheet(ws128, results, bit=128)

    ws256 = wb.create_sheet("OT-256 Benchmark")
    _write_ot_sheet(ws256, results, bit=256)

    ws_plain = wb.create_sheet("Plain Intersection")
    _write_plain_sheet(ws_plain, results)

    wb.save(out_path)
    n128  = len([r for r in results if r.mode == "OT-128"])
    n256  = len([r for r in results if r.mode == "OT-256"])
    nplain = len([r for r in results if r.mode == "PLAIN"])
    print(f"\nExcel hazir: {out_path}")
    print(f"  Sheet 1: OT-128 Benchmark   - {n128} satir")
    print(f"  Sheet 2: OT-256 Benchmark   - {n256} satir")
    print(f"  Sheet 3: Plain Intersection - {nplain} satir (bit sutunu yok)")

# ---------------------------------------------------------------------------
# BENCHMARK CORE (sadece kripto/matematik, LLM yok)
# ---------------------------------------------------------------------------

def _time_plain(cfg: ConfigRow) -> RunResult:
    """Plain intersection'i olcer, LLM calistirmaz."""
    t0 = time.perf_counter()
    try:
        arr = plain_intersection(cfg.agent_a, cfg.agent_b)
        duration = time.perf_counter() - t0
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=0,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode="PLAIN",
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=duration,
            intersection_array=str(list(arr)),
            intersection_size=len(arr),
            success=True,
        )
    except Exception as e:
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=0,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode="PLAIN",
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=time.perf_counter() - t0,
            error=str(e),
        )


def _time_ot(cfg: ConfigRow, bit_size: int) -> RunResult:
    """OT intersection'i olcer (LLM yok)."""
    mode = f"OT-{bit_size}"
    t0 = time.perf_counter()
    try:
        mgr = OTManager(bit_size=bit_size)
        arr = mgr.compute_intersection(cfg.agent_a, cfg.agent_b, len(cfg.agent_a))
        duration = time.perf_counter() - t0
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=bit_size,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode=mode,
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=duration,
            intersection_array=str(arr),
            intersection_size=len(arr),
            success=True,
        )
    except Exception as e:
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=bit_size,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode=mode,
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=time.perf_counter() - t0,
            error=str(e),
        )

# ---------------------------------------------------------------------------
# E2E LLM RUNNER  (eger Terrarium modulleri mevcutsa)
# ---------------------------------------------------------------------------

async def _run_e2e_for_mode(
    cfg: ConfigRow,
    base_config: dict,
    mode: str,       # "OT-128" | "OT-256" | "PLAIN"
    bit_size: int,
) -> RunResult:
    """Tek bir config x mod icin tam LLM simülasyonunu calistirir."""
    _patch_environment(cfg.agent_a, cfg.agent_b)

    if mode == "PLAIN":
        _patch_crypto("PLAIN", 0)
    else:
        _patch_crypto("OT", bit_size)

    run_cfg = dict(base_config)
    run_cfg["simulation"] = dict(base_config.get("simulation", {}))
    run_cfg["simulation"]["run_timestamp"] = (
        f"benchmark_cfg{cfg.config_id}_{mode.lower().replace('-','')}"
    )
    run_cfg["simulation"]["note"] = f"E2E Benchmark - Config {cfg.config_id} - {mode}"
    run_cfg["environment"] = dict(base_config.get("environment", {}))
    run_cfg["environment"]["num_days"] = cfg.negotiation_window_days
    run_cfg["environment"]["slots_per_day"] = (
        cfg.total_slots // max(cfg.negotiation_window_days, 1)
    )

    t0 = time.perf_counter()
    try:
        success = await run_simulation(run_cfg)
        duration = time.perf_counter() - t0
        # E2E'de intersection array'i dogrudan hesaplayarak kaydedelim
        arr = plain_intersection(cfg.agent_a, cfg.agent_b)
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=bit_size,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode=mode,
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=duration,
            intersection_array=str(list(arr)),
            intersection_size=len(arr),
            success=bool(success),
        )
    except Exception as e:
        return RunResult(
            config_id=cfg.config_id,
            total_slots=cfg.total_slots,
            security_bits=bit_size,
            availability_density=cfg.availability_density,
            intersection_density=cfg.intersection_density,
            actual_intersections=cfg.actual_intersections,
            mode=mode,
            agent_a=cfg.agent_a,
            agent_b=cfg.agent_b,
            duration_s=time.perf_counter() - t0,
            error=str(e),
        )

# ---------------------------------------------------------------------------
# MAIN ORCHESTRATOR
# ---------------------------------------------------------------------------

async def run_benchmark(
    csv_path: Path,
    config_path: Optional[Path],
    out_path: Path,
    max_rows: Optional[int],
    no_ot: bool,
    e2e_mode: bool,
) -> None:
    """
    Ana orkestrator:
      ASAMA 1: OT-128 (tum config'ler)
      ASAMA 2: OT-256 (tum config'ler)
      ASAMA 3: Plain  (tum config'ler)
    """
    configs = load_configurations(csv_path)
    if not configs:
        print("HATA: Hic konfigürasyon yuklenemedi.")
        return
    if max_rows:
        configs = configs[:max_rows]

    run_ot = OT_AVAILABLE and not no_ot

    print(f"\n{'='*60}")
    print(f"  E2E Benchmark -- 3 Asama, 3 Sheet Excel")
    print(f"  Konfigurasyon : {len(configs)}")
    print(f"  OT aktif      : {run_ot}")
    print(f"  E2E LLM modu  : {e2e_mode and E2E_AVAILABLE}")
    print(f"  Cikti         : {out_path}")
    print(f"{'='*60}")

    base_config = None
    if e2e_mode and E2E_AVAILABLE and config_path:
        from src.utils import load_config  # type: ignore
        base_config = load_config(str(config_path))
        print(f"  LLM Modeli    : {base_config.get('llm', {}).get('model', '?')}")

    all_results: List[RunResult] = []
    n = len(configs)

    # ----------------------------------------------------------------
    # ASAMA 1: OT-128
    # ----------------------------------------------------------------
    if run_ot:
        print(f"\n{'-'*60}")
        print(f"  ASAMA 1/3: OT-128 ({n} config)")
        print(f"{'-'*60}")
        for i, cfg in enumerate(configs, 1):
            print(f"  [{i:3d}/{n}] config={cfg.config_id} | slots={cfg.total_slots}", end=" ... ", flush=True)
            if e2e_mode and E2E_AVAILABLE and base_config:
                r = await _run_e2e_for_mode(cfg, base_config, "OT-128", 128)
            else:
                r = _time_ot(cfg, 128)
            all_results.append(r)
            if r.error:
                print(f"{r.duration_s:.6f} s | HATA: {r.error[:40]}")
            else:
                print(f"{r.duration_s:.6f} s | boyut={r.intersection_size}")

    # ----------------------------------------------------------------
    # ASAMA 2: OT-256
    # ----------------------------------------------------------------
    if run_ot:
        print(f"\n{'-'*60}")
        print(f"  ASAMA 2/3: OT-256 ({n} config)")
        print(f"{'-'*60}")
        for i, cfg in enumerate(configs, 1):
            print(f"  [{i:3d}/{n}] config={cfg.config_id} | slots={cfg.total_slots}", end=" ... ", flush=True)
            if e2e_mode and E2E_AVAILABLE and base_config:
                r = await _run_e2e_for_mode(cfg, base_config, "OT-256", 256)
            else:
                r = _time_ot(cfg, 256)
            all_results.append(r)
            if r.error:
                print(f"{r.duration_s:.6f} s | HATA: {r.error[:40]}")
            else:
                print(f"{r.duration_s:.6f} s | boyut={r.intersection_size}")

    if not run_ot:
        print("\n  OT atlandi (--no-ot veya pyot yuklu degil)")

    # ----------------------------------------------------------------
    # ASAMA 3: PLAIN
    # ----------------------------------------------------------------
    print(f"\n{'-'*60}")
    print(f"  ASAMA {'3' if run_ot else '1'}/{'3' if run_ot else '1'}: Plain Intersection ({n} config)")
    print(f"{'-'*60}")
    for i, cfg in enumerate(configs, 1):
        print(f"  [{i:3d}/{n}] config={cfg.config_id} | slots={cfg.total_slots}", end=" ... ", flush=True)
        if e2e_mode and E2E_AVAILABLE and base_config:
            r = await _run_e2e_for_mode(cfg, base_config, "PLAIN", 0)
        else:
            r = _time_plain(cfg)
        all_results.append(r)
        print(f"{r.duration_s:.6f} s | boyut={r.intersection_size}")

    # ----------------------------------------------------------------
    # EXCEL YAZ
    # ----------------------------------------------------------------
    write_excel(all_results, out_path)

    # Ozet
    print(f"\n{'='*60}")
    print("  OZET")
    print(f"{'='*60}")
    for mode in (["OT-128", "OT-256"] if run_ot else []) + ["PLAIN"]:
        mrs = [r for r in all_results if r.mode == mode]
        if mrs:
            ok = sum(1 for r in mrs if not r.error)
            avg = mean(r.duration_s for r in mrs)
            avg_size = mean(r.intersection_size for r in mrs)
            print(f"  {mode:<12}: {len(mrs)} config | {ok} basarili | ort. sure {avg:.6f} s | ort. boyut {avg_size:.2f}")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="E2E Benchmark — OT-128, OT-256, Plain Intersection → 3 Sheet Excel",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--csv",      type=Path, default=DEFAULT_CSV,
                   help="CSV dosya yolu")
    p.add_argument("--config",   type=Path,
                   default=REPO_ROOT / "examples/configs/meeting_scheduling.yaml",
                   help="Terrarium YAML config yolu (E2E modunda kullanilir)")
    p.add_argument("--out",      type=Path, default=DEFAULT_OUT,
                   help="Cikti Excel dosyasi")
    p.add_argument("--max-rows", type=int,  default=None,
                   help="Sinirli sayida config calistir (test icin)")
    p.add_argument("--no-ot",    action="store_true",
                   help="OT benchmark'larini atla, sadece Plain calistir")
    p.add_argument("--e2e",      action="store_true",
                   help="Tam LLM simulasyonu calistir (varsayilan: sadece kripto zamanlama)")
    p.add_argument("--log-level", default="WARNING",
                   choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    logging.getLogger().setLevel(getattr(logging, args.log_level))
    (TESTS_DIR / "results").mkdir(parents=True, exist_ok=True)
    asyncio.run(run_benchmark(
        csv_path=args.csv,
        config_path=args.config,
        out_path=args.out,
        max_rows=args.max_rows,
        no_ot=args.no_ot,
        e2e_mode=args.e2e,
    ))